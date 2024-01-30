from unittest import TestCase
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


def get_cells(ws: Worksheet) -> list[list[any]]:
    result = []
    for row in ws.rows:
        new_row = []
        for cell in row:
            value = str(cell.value).strip()
            if value != '':
                new_row.append(value)
        result.append(new_row)
    return result


class TestNormalForm1(TestCase):
    def setUp(self):
        self.wb = load_workbook('table1.xlsx')

    def test_workbook_has_correct_sheets(self):
        self.assertIsNotNone(self.wb)
        sheet_names = self.wb.sheetnames
        self.assertEqual(len(sheet_names), 2)
        self.assertIn('persons', sheet_names)
        self.assertIn('grades', sheet_names)

    def test_persons_sheet_is_correct(self):
        cells = get_cells(self.wb['persons'])
        columns = [len(row) for row in cells]
        self.assertEqual(len(columns), 4)
        self.assertTrue(all(map(lambda x: x == 3, columns)))
        values = [cell for row in cells for cell in row]
        self.assertEqual(values.count('Иванов'), 1)
        self.assertEqual(values.count('Иван'), 2)
        self.assertEqual(values.count('Иванович'), 1)
        self.assertEqual(values.count('Захаров'), 1)
        self.assertEqual(values.count('Егорович'), 1)
        self.assertEqual(values.count('Винокуров'), 1)
        self.assertEqual(values.count('Александр'), 1)
        self.assertEqual(values.count('Романович'), 1)

    def test_grades_sheet_is_correct(self):
        cells = get_cells(self.wb['grades'])
        columns = [len(row) for row in cells]
        self.assertEqual(len(columns), 6)
        self.assertTrue(all(map(lambda x: x == 4, columns)))
        values = [tuple(row) for row in cells]
        self.assertIn(('2', 'ФИИТ-22', 'Информатика', '5'), values)
        self.assertIn(('2', 'ФИИТ-22', 'Математика', '4'), values)
        self.assertIn(('2', 'ФИИТ-22', 'Базы данных', '4'), values)
        self.assertIn(('5', 'ИВТ-22', 'Базы данных', '5'), values)
        self.assertIn(('5', 'ИВТ-22', 'Математика', '4'), values)


class TestNormalForm2(TestCase):
    def setUp(self):
        self.wb = load_workbook('table2.xlsx')

    def test_workbook_has_correct_sheets(self):
        self.assertIsNotNone(self.wb)
        sheet_names = self.wb.sheetnames
        self.assertEqual(len(sheet_names), 2)
        self.assertIn('employee', sheet_names)
        self.assertIn('students', sheet_names)

    def test_employee_sheet_is_correct(self):
        cells = get_cells(self.wb['employee'])
        columns = [len(row) for row in cells]
        self.assertEqual(len(columns), 9)
        self.assertTrue(all(map(lambda x: x == 6, columns)))
        values = [cell for row in cells for cell in row]
        text = ' '.join(values)
        self.assertEqual(text.count('(ключ)'), 1)

    def test_students_sheet_is_correct(self):
        cells = get_cells(self.wb['students'])
        columns = [len(row) for row in cells]
        self.assertEqual(len(columns), 9)
        self.assertTrue(all(map(lambda x: x == 6, columns)))
        values = [cell for row in cells for cell in row]
        words = [
            'Александр', 'Александра', 'Александровна', 'Анастасия', 'Андрей',
            'Аркадьевич', 'Артамонова', 'Артём', 'Беляев', 'Василий',
            'Винокуров', 'Дамирович', 'Евгений', 'Илларионов', 'Ильинична',
            'Ильич', 'Марковна', 'Митрофанова', 'Михайлович', 'Назаров',
            'Романович', 'Рыбакова', 'Тарасов', 'Татьяна',
        ]
        for word in words:
            self.assertEqual(values.count(word), 1)
        text = ' '.join(values)
        self.assertEqual(text.count('(ключ)'), 1)


class TestNormalForm3(TestCase):
    def setUp(self):
        self.wb = load_workbook('table3.xlsx')

    def test_workbook_has_correct_sheets(self):
        self.assertIsNotNone(self.wb)
        sheet_names = self.wb.sheetnames
        self.assertEqual(len(sheet_names), 5)
        self.assertIn('employee', sheet_names)
        self.assertIn('specialties', sheet_names)
        self.assertIn('departments', sheet_names)
        self.assertIn('faculties', sheet_names)
        self.assertIn('levels', sheet_names)

    # def test_employee_sheet_is_correct(self):
    #     cells = get_cells(self.wb['employee'])
    #     columns = [len(row) for row in cells]
    #     self.assertEqual(len(columns), 11)
    #     self.assertTrue(all(map(lambda x: x == 3, columns)))
    #     values = [cell for row in cells for cell in row]
    #     for val in self.positions:
    #         self.assertNotIn(val, values)
    #     for val in self.employee:
    #         self.assertIn(val, values)
    #
    # def test_positions_sheet_is_correct(self):
    #     cells = get_cells(self.wb['positions'])
    #     columns = [len(row) for row in cells]
    #     self.assertEqual(len(columns), 6)
    #     self.assertTrue(all(map(lambda x: x == 3, columns)))
    #     values = [cell for row in cells for cell in row]
    #     for val in self.positions:
    #         self.assertIn(val, values)
    #     for val in self.employee:
    #         self.assertNotIn(val, values)
    #
    # def test_students_sheet_is_correct(self):
    #     cells = get_cells(self.wb['students'])
    #     columns = [len(row) for row in cells]
    #     self.assertEqual(len(columns), 16)
    #     self.assertTrue(all(map(lambda x: x == 5, columns)))
    #     values = [cell for row in cells for cell in row]
    #     for val in self.students:
    #         self.assertIn(val, values)
    #     for val in self.specialties:
    #         self.assertNotIn(val, values)
    #
    # def test_specialty_sheet_is_correct(self):
    #     cells = get_cells(self.wb['specialty'])
    #     columns = [len(row) for row in cells]
    #     self.assertEqual(len(columns), 6)
    #     self.assertTrue(all(map(lambda x: x == 3, columns)))
    #     values = [cell for row in cells for cell in row]
    #     for val in self.students:
    #         self.assertNotIn(val, values)
    #     for val in self.specialties:
    #         self.assertIn(val, values)
